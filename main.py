import openpyxl
import os

spreadsheet = openpyxl.load_workbook("produceSales.xlsx")


sheet = spreadsheet[spreadsheet.sheetnames[0]]

producesUpdate = {
    "Potatoes": 10,
    "Okra": 15,
    "Fava Beans": 20,
    "Watermelon": 30,
}


for row in range(2, sheet.max_row):
    nameProduce = sheet.cell(row=row, column=1).value
    if nameProduce in producesUpdate:
        oldValue = sheet.cell(row=row, column=2).value
        sheet.cell(row=row, column=2).value = producesUpdate[nameProduce]
        print(nameProduce + " : " + str(oldValue) +
              " => "+str(producesUpdate[nameProduce]))
spreadsheet.save("updated.xlsx")
